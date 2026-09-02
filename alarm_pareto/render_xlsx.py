"""Step 5: write the Excel workbook with native charts.

The charts here are real Excel chart objects, not pictures. You can click into
them in Excel 2016, re-sort the source table, and the chart updates. This is why
we use openpyxl chart objects instead of pasting images.

Each summary sheet shows a Pareto twice. Once ranked by how often a fault
happens. Once ranked by downtime. The two are shown side by side so you never
have to open a second file to compare them.
"""

from datetime import datetime

import numpy as np
import pandas as pd
from openpyxl import Workbook
from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from . import aggregate as agg

# Simple styles reused across sheets.
HEADER_FONT = Font(bold=True, color="FFFFFF")
HEADER_FILL = PatternFill("solid", fgColor="305496")
TITLE_FONT = Font(bold=True, size=13)
NOTE_FONT = Font(italic=True, size=10, color="555555")

# Number formats. These control how Excel displays a cell, not the stored value.
FMT_DATETIME = "yyyy-mm-dd hh:mm:ss"
FMT_HOURS = "0.00"
FMT_PCT = "0.0"


def write_workbook(result, windowed_table, output_path):
    """Write the full workbook to output_path.

    result: the dictionary returned by aggregate.aggregate.
    windowed_table: the filtered rows for the window (the raw data sheet).
    output_path: where to save the .xlsx file.
    """
    wb = Workbook()
    # A new workbook starts with one empty sheet. We reuse it as the data sheet.
    data_ws = wb.active
    data_ws.title = "Window_Data"
    _write_window_data(data_ws, windowed_table, result)

    method = result["downtime_method"]
    method_label = _method_label(method)

    # One summary sheet per grouping level.
    sheet_names = {
        "fault_code": "By_Fault_Code",
        "description": "By_Description",
        "equipment": "By_Equipment",
    }
    for level in agg.GROUPING_LEVELS:
        ws = wb.create_sheet(sheet_names[level])
        _write_summary_sheet(ws, result, level, method, method_label)

    wb.save(output_path)
    return output_path


def _method_label(method):
    if method == agg.METHOD_ATTRIBUTED:
        return "Attributed downtime (each fault credited its full duration)"
    if method == agg.METHOD_IN_RANGE:
        return ("In-range downtime (overlaps merged, and each fault cut down to "
                "the hours this report covers)")
    return "True wall-clock downtime (overlaps merged, counted once)"


def _write_window_data(ws, windowed_table, result):
    """Write the filtered raw rows for the window onto the data sheet."""
    grand = result["grand"]
    ws["A1"] = "Filtered alarm data for the reporting window"
    ws["A1"].font = TITLE_FONT
    ws["A2"] = "Window start: %s   Window end: %s   Length: %s days   Time of day: %s" % (
        _fmt_ts(grand["window_start"]),
        _fmt_ts(grand["window_end"]),
        grand["window_days"],
        grand["time_of_day_label"],
    )
    ws["A2"].font = NOTE_FONT

    # Spell out the three downtime numbers here, on the data sheet, where the
    # rows they came from are sitting right underneath.
    ws["A3"] = (
        "Downtime, three ways. Attributed %.2f h (each fault its full duration). "
        "True wall clock %.2f h (overlaps merged). "
        "In range %.2f h of the %.2f h this report covers, %.1f%% (overlaps merged and "
        "each fault cut to the covered hours). The three answer different questions. "
        "Do not mix them." % (
            grand["attributed_downtime_hours"],
            grand["wallclock_downtime_hours"],
            grand["in_range_downtime_hours"],
            grand["range_hours"],
            grand["in_range_downtime_pct"],
        )
    )
    ws["A3"].font = NOTE_FONT
    ws["A4"] = "Time covered: %s" % grand["range_description"]
    ws["A4"].font = NOTE_FONT

    start_row = 6
    _write_dataframe(ws, windowed_table, start_row=start_row, start_col=1)
    _autosize(ws, windowed_table, start_col=1)


def _write_summary_sheet(ws, result, level, method, method_label):
    """Write the count Pareto and the downtime Pareto for one grouping level."""
    tables = result["levels"][level]
    by_count = tables["by_count"]
    by_downtime = tables["by_downtime"]

    nice_level = {"fault_code": "Fault Code", "description": "Description", "equipment": "Equipment"}[level]

    ws["A1"] = "Pareto by %s" % nice_level
    ws["A1"].font = TITLE_FONT
    ws["A2"] = "Downtime method used for the downtime ranking: %s" % method_label
    ws["A2"].font = NOTE_FONT
    ws["A3"] = "Top %d rows are shown. The rest are grouped as 'Other'." % result["top_n"]
    ws["A3"].font = NOTE_FONT

    table_start_row = 5
    # Both tables have the same number of rows, so they end on the same row.
    n_rows = len(by_count)
    tables_bottom = table_start_row + 1 + n_rows  # header row plus data rows

    # The two charts are stacked one above the other, both anchored at column A.
    # Stacking avoids the charts overlapping, which would happen if we placed
    # them side by side because each chart is wider than its table.
    count_chart_anchor = "A%d" % (tables_bottom + 2)
    downtime_chart_anchor = "A%d" % (tables_bottom + 26)

    # Left table: ranked by occurrence count.
    count_display = _select_columns(
        by_count, level,
        [("rank", "Rank"), (level, nice_level), ("count", "Count"),
         ("count_pct", "Count %"), ("cum_count_pct", "Cumulative %")],
    )
    _write_block(
        ws, count_display, title="Ranked by occurrence count",
        start_row=table_start_row, start_col=1,
        category_display_col=2, bar_display_col=3, cum_display_col=5,
        bar_axis_title="Count", pct_cols=[4, 5], hours_cols=[],
        chart_anchor=count_chart_anchor,
    )

    # Right table: ranked by downtime, using the chosen method.
    metric_hours = agg.METHOD_COLUMNS[method].replace("_s", "_hours")
    downtime_display = _select_columns(
        by_downtime, level,
        [("rank", "Rank"), (level, nice_level), ("count", "Count"),
         (metric_hours, "Downtime (h)"), ("downtime_pct", "Downtime %"),
         ("cum_downtime_pct", "Cumulative %")],
    )
    _write_block(
        ws, downtime_display, title="Ranked by downtime (%s)" % method,
        start_row=table_start_row, start_col=8,
        category_display_col=2, bar_display_col=4, cum_display_col=6,
        bar_axis_title="Downtime (hours)", pct_cols=[5, 6], hours_cols=[4],
        chart_anchor=downtime_chart_anchor,
    )


def _select_columns(df, level, pairs):
    """Pick and rename columns for display. pairs is a list of (source, label)."""
    if df.empty:
        return pd.DataFrame(columns=[label for _, label in pairs])
    out = pd.DataFrame()
    for source, label in pairs:
        out[label] = df[source].values
    return out


def _write_block(ws, display_df, title, start_row, start_col,
                 category_display_col, bar_display_col, cum_display_col,
                 bar_axis_title, pct_cols, hours_cols, chart_anchor):
    """Write one table plus its native Pareto chart at the given chart anchor.

    category_display_col, bar_display_col, cum_display_col are 1-based column
    numbers within the displayed table. They tell the chart which columns hold
    the category labels, the bar values, and the cumulative percent line.
    """
    ws.cell(row=start_row, column=start_col, value=title).font = Font(bold=True, size=11)

    header_row = start_row + 1
    n_cols = display_df.shape[1]

    # Add a hidden helper column holding the value 80 on every data row. The
    # Pareto reference line is drawn from this column. openpyxl has no built-in
    # horizontal line, so a flat data series is the normal way to draw one.
    ref_col_index = start_col + n_cols  # one past the last displayed column

    # Write the table headers.
    for j, col_name in enumerate(display_df.columns):
        c = ws.cell(row=header_row, column=start_col + j, value=str(col_name))
        c.font = HEADER_FONT
        c.fill = HEADER_FILL
        c.alignment = Alignment(horizontal="center")
    ws.cell(row=header_row, column=ref_col_index, value="Ref80").font = HEADER_FONT
    ws.cell(row=header_row, column=ref_col_index).fill = HEADER_FILL

    # Write the data rows.
    n_rows = len(display_df)
    for i in range(n_rows):
        for j in range(n_cols):
            value = display_df.iloc[i, j]
            cell = ws.cell(row=header_row + 1 + i, column=start_col + j, value=_clean(value))
            col_1_based = j + 1
            if col_1_based in pct_cols:
                cell.number_format = FMT_PCT
            elif col_1_based in hours_cols:
                cell.number_format = FMT_HOURS
        ws.cell(row=header_row + 1 + i, column=ref_col_index, value=80).number_format = FMT_PCT

    if n_rows == 0:
        ws.cell(row=header_row + 1, column=start_col, value="No data in the window.")
        return

    first_data_row = header_row + 1
    last_data_row = header_row + n_rows

    # Build the chart. Bars for the primary value. A line for cumulative percent
    # on a second axis. A flat line at 80 percent for the reference.
    bar = BarChart()
    bar.type = "col"
    bar.title = title
    bar.y_axis.title = bar_axis_title
    bar.x_axis.title = str(display_df.columns[category_display_col - 1])

    bar_value_ref = Reference(
        ws,
        min_col=start_col + bar_display_col - 1,
        min_row=header_row,
        max_row=last_data_row,
    )
    cats_ref = Reference(
        ws,
        min_col=start_col + category_display_col - 1,
        min_row=first_data_row,
        max_row=last_data_row,
    )
    bar.add_data(bar_value_ref, titles_from_data=True)
    bar.set_categories(cats_ref)

    line = LineChart()
    cum_ref = Reference(
        ws,
        min_col=start_col + cum_display_col - 1,
        min_row=header_row,
        max_row=last_data_row,
    )
    ref80_ref = Reference(ws, min_col=ref_col_index, min_row=header_row, max_row=last_data_row)
    line.add_data(cum_ref, titles_from_data=True)
    line.add_data(ref80_ref, titles_from_data=True)

    # Put the line on a secondary axis fixed to 0..100 percent.
    line.y_axis.axId = 200
    line.y_axis.title = "Cumulative %"
    line.y_axis.scaling.min = 0
    line.y_axis.scaling.max = 100
    # crosses="max" moves the second axis to the right side of the chart.
    line.y_axis.crosses = "max"

    # Combining the two charts overlays the lines on the bars.
    bar += line

    bar.width = 20
    bar.height = 11

    # Place the chart at the anchor the caller chose. Anchors are stacked so the
    # two charts on a sheet do not overlap.
    ws.add_chart(bar, chart_anchor)


def _write_dataframe(ws, df, start_row, start_col):
    """Write a DataFrame with a styled header row. Handles dates and blanks."""
    for j, col_name in enumerate(df.columns):
        c = ws.cell(row=start_row, column=start_col + j, value=str(col_name))
        c.font = HEADER_FONT
        c.fill = HEADER_FILL
    for i in range(len(df)):
        for j, col_name in enumerate(df.columns):
            value = df.iloc[i, j]
            cell = ws.cell(row=start_row + 1 + i, column=start_col + j, value=_clean(value))
            if isinstance(_clean(value), datetime):
                cell.number_format = FMT_DATETIME


def _clean(value):
    """Convert pandas and numpy values into things openpyxl can store safely.

    openpyxl wants plain Python types. pandas and numpy hand back their own
    number and date types, so we translate them here. Missing values become
    blank cells.
    """
    if value is None:
        return None
    if value is pd.NaT:
        return None
    # numpy has its own integer and float types. Turn them into plain Python.
    if isinstance(value, np.integer):
        return int(value)
    if isinstance(value, np.floating):
        return None if np.isnan(value) else float(value)
    # pandas uses NaN for missing numbers and NaT for missing dates.
    if isinstance(value, float) and pd.isna(value):
        return None
    if isinstance(value, pd.Timestamp):
        return None if pd.isna(value) else value.to_pydatetime()
    return value


def _fmt_ts(ts):
    if ts is None or (isinstance(ts, pd.Timestamp) and pd.isna(ts)):
        return "n/a"
    return pd.Timestamp(ts).strftime("%Y-%m-%d %H:%M:%S")


def _autosize(ws, df, start_col):
    """Rough column width based on the header text length."""
    for j, col_name in enumerate(df.columns):
        letter = get_column_letter(start_col + j)
        ws.column_dimensions[letter].width = max(12, min(40, len(str(col_name)) + 4))
